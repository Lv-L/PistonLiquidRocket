"""
Baseline check: PistonLiquidRocket vs HalfCatSim_v1.3.8_MojaveSphinx - R01.xlsx

Reference values taken directly from the Motor Simulation sheet of the spreadsheet.
Run with:  python baseline_mojave.py
"""

import sys, os, time
import numpy as np
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from PistonLiquidRocket import (
    EngineConfig, PistonEngine, Fuel, FUELS,
    print_config, print_summary, G0, N2O_FLUID,
    _ROCKETCEA_AVAILABLE,
)
from CoolProp.CoolProp import PropsSI as _CP

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ══════════════════════════════════════════════════════════════════════════════
#  HalfCatSim reference output values  (from Motor Simulation sheet)
# ══════════════════════════════════════════════════════════════════════════════
REF = {
    'pc_mpa':       255.46 / 145.038,        # 1.761 MPa
    'thrust_N':     258.41 * 4.44822,        # 1149.3 N
    'burn_s':       5.31,
    'of':           2.10,
    'mdot_kgs':     1.1286,
    'isp_s':        98.2,                    # corrected Isp (with etas)
    'J_Ns':         4487.9,
    'ox_mass_kg':   3.108,
    'fuel_mass_kg': 1.551,
    'pe_bar':       0.675,
}

# ══════════════════════════════════════════════════════════════════════════════
#  Derived constants (all converted to SI)
# ══════════════════════════════════════════════════════════════════════════════
IN = 0.0254   # inch -> metre

# Tank geometry: OD=4 in, ID=3.75 in, Ox=20 in, Fuel=11 in
TANK_ID_M    = 3.75 * IN
OX_LENGTH_M  = 20.0 * IN          # 508 mm
FUEL_LENGTH_M= 11.0 * IN          # 279.4 mm
_A_tank      = np.pi / 4.0 * TANK_ID_M**2
V_OX_GEOM    = _A_tank * OX_LENGTH_M    # 3.620e-3 m³
V_FUEL_GEOM  = _A_tank * FUEL_LENGTH_M  # 1.991e-3 m³
TANK_VOL     = V_OX_GEOM + V_FUEL_GEOM  # 5.611e-3 m³

# Cv=1.6 ball valve  →  CdA (water, 60°F basis)
#   CdA [m²] = Cv × 1.6984e-5   (derived from Cv definition)
CV = 1.6
VALVE_CDA  = CV * 1.6984e-5           # 2.717e-5 m²
VALVE_DIAM = np.sqrt(4 * VALVE_CDA / np.pi)   # 5.882 mm  (used with Cd=1.0)

# Injector holes
OX_HOLE_D   = 0.0995 * IN             # 2.5273 mm
FUEL_HOLE_D = 0.0748 * IN             # 1.8999 mm
A_INJ_OX    = 6 * np.pi / 4 * OX_HOLE_D**2    # 6 holes: 30.10e-6 m²
A_INJ_FUEL  = 6 * np.pi / 4 * FUEL_HOLE_D**2  # 6 holes: 17.01e-6 m²

# Ullage: to match ox_mass=3.108 kg in V_ox=3.620e-3 m³ almost exactly.
# At 282.04 K: rho_liq=858.73, rho_vap=111.09  → ullage ≈ 0.015% (essentially 0).
# Using 0.05% gives ox_mass within ~0.05% of target.
OX_ULLAGE = 0.0005

# E85 with HalfCatSim density.
# Motor Simulation!K3 = 'E85', Fuel!density = 774 kg/m³.
E85_HCS = Fuel(
    name='E85', cea_name=FUELS['E85'].cea_name,
    cea_card=FUELS['E85'].cea_card,
    coolprop_name=None,                # bypass CoolProp to force density_ref
    density_ref=774.0,                 # HalfCatSim fuel density
    mw_g_mol=FUELS['E85'].mw_g_mol,
    n_o2_stoich=FUELS['E85'].n_o2_stoich,
    t_ad_peak=FUELS['E85'].t_ad_peak,
    t_ad_of_peak=FUELS['E85'].t_ad_of_peak,
    t_ad_sigma=FUELS['E85'].t_ad_sigma,
    t_ad_floor=FUELS['E85'].t_ad_floor,
    mw_exh_a=FUELS['E85'].mw_exh_a,
    mw_exh_b=FUELS['E85'].mw_exh_b,
    gamma_a=FUELS['E85'].gamma_a,
    gamma_b=FUELS['E85'].gamma_b,
)

# ══════════════════════════════════════════════════════════════════════════════
#  Engine configuration  — matched to HalfCatSim inputs
# ══════════════════════════════════════════════════════════════════════════════
cfg = EngineConfig(
    # --- Propellant / operating point ---
    fuel             = E85_HCS,
    n2o_init_temp_K  = 282.04,         # 48°F
    of_ratio         = 2.10,
    ox_ullage        = OX_ULLAGE,

    # --- Tank geometry ---
    ox_tank_diam     = TANK_ID_M,      # 95.25 mm (3.75 in ID)
    fuel_tank_diam   = TANK_ID_M,
    tank_volume      = TANK_VOL,       # 5.611 L (ox 20 in + fuel 11 in)

    # --- Injectors (manual, from hole geometry) ---
    A_inj_ox         = A_INJ_OX,       # 6 × 0.0995 in
    cd_ox            = 0.48,
    A_inj_fuel       = A_INJ_FUEL,     # 6 × 0.0748 in
    cd_fuel          = 0.48,

    # --- Oxidiser feed line (0.3125 in ID, 12 in, roughness 3.94e-6 in) ---
    ox_line_id        = 0.3125 * IN,
    ox_line_length    = 12.0  * IN,
    ox_line_roughness = 3.94e-6 * IN,   # 1.001e-7 m (drawn tube)
    ox_valve_Cd       = 1.0,
    ox_valve_diam     = VALVE_DIAM,     # Cv=1.6

    # --- Fuel feed line (0.225 in ID, 48 in, roughness 0.0002 in) ---
    fuel_line_id        = 0.225 * IN,
    fuel_line_length    = 48.0  * IN,
    fuel_line_roughness = 0.0002 * IN,  # 5.08e-6 m (flexible hose)
    fuel_valve_Cd       = 1.0,
    fuel_valve_diam     = VALVE_DIAM,   # Cv=1.6

    # --- Combustion efficiency ---
    cstar_eta        = 0.60,            # HalfCatSim
    nozzle_eta       = 0.97,

    # --- Chamber geometry ---
    chamber_diam          = 2.0 * IN,   # 50.8 mm
    chamber_length        = 5.0 * IN,   # 127 mm
    convergent_half_angle = 45.0,

    # --- Nozzle ---
    throat_diam      = 1.0 * IN,        # 25.4 mm
    AeAt             = 4.0,

    # --- Environment ---
    p_ambient        = 94_197.0,        # Pa  (2000 ft altitude)

    # --- Piston pressurant (N2O vapour space) ---
    ipa_uses_n2o_p   = True,
    piston_dp        = 15 * 6894.76,    # 15 psi = 103,421 Pa
)

# ══════════════════════════════════════════════════════════════════════════════
#  Run
# ══════════════════════════════════════════════════════════════════════════════
if not _ROCKETCEA_AVAILABLE:
    print("  WARNING: rocketcea not found -- using polynomial CEA fits.")

eng = PistonEngine(cfg)
print_config(cfg, eng)

print(f"\n  Simulating ({E85_HCS.name}, adiabatic N2O, dt=0.005 s)...")
t0   = time.perf_counter()
hist = eng.run(dt=0.005)
print(f"  Done in {time.perf_counter()-t0:.2f} s  ({len(hist['t'])} steps)")
print_summary(hist, eng.stop_reason)

# ══════════════════════════════════════════════════════════════════════════════
#  Comparison table
# ══════════════════════════════════════════════════════════════════════════════
if not hist['t']:
    sys.exit(0)

t_arr  = np.array(hist['t'])
thr    = np.array(hist['thrust'])
isp    = np.array(hist['isp'])
mdot   = np.array(hist['mdot'])
p_c    = np.array(hist['p_chamber'])
J      = float(np.trapezoid(thr, t_arr))
t_burn = t_arr[-1]

# Initial-condition values (first step) and burn-time averages
sim = {
    'pc_mpa':       p_c[0] / 1e6,
    'thrust_N':     thr[0],
    'burn_s':       t_burn,
    'of':           float(np.mean(hist['of_actual'])),
    'mdot_kgs':     float(mdot[0]),
    'isp_s':        float(np.mean(isp)),
    'J_Ns':         J,
    'ox_mass_kg':   cfg.ox_mass,
    'fuel_mass_kg': eng.fuel_tank.propellant_mass,
    'pe_bar':       eng._pe_over_pc * p_c[0] / 1e5,
}

ROWS = [
    ('pc_mpa',       'Chamber pressure (MPa)',     '{:.3f}', 'initial'),
    ('thrust_N',     'Thrust (N)',                  '{:.1f}', 'initial'),
    ('burn_s',       'Burn time (s)',               '{:.2f}', ''),
    ('of',           'O/F ratio',                   '{:.3f}', 'burn avg'),
    ('mdot_kgs',     'Mass flow (kg/s)',             '{:.4f}', 'initial'),
    ('isp_s',        'Isp (s)',                      '{:.1f}', 'burn avg'),
    ('J_Ns',         'Total impulse (N·s)',          '{:.0f}', ''),
    ('ox_mass_kg',   'N2O loaded (kg)',              '{:.3f}', ''),
    ('fuel_mass_kg', 'Fuel loaded (kg)',             '{:.3f}', ''),
    ('pe_bar',       'Nozzle exit pressure (bar)',   '{:.3f}', 'initial'),
]

W = 65
print()
print('=' * W)
print('  Baseline: PistonLiquidRocket  vs  HalfCatSim v1.3.8  (Mojave Sphinx R01)')
print('=' * W)
print(f"  {'Parameter':<29} {'Sim':>9}  {'HalfCat':>9}  {'Err %':>7}  Note")
print('-' * W)
for key, label, fmt, note in ROWS:
    s = sim[key];  r = REF[key]
    err = (s - r) / (abs(r) + 1e-30) * 100
    note_str = f'  {note}' if note else ''
    print(f"  {label:<29} {fmt.format(s):>9}  {fmt.format(r):>9}  {err:>+6.1f}%{note_str}")
print('=' * W)

print()
print('  HalfCatSim is a steady-state model; simulation is time-varying (adiabatic N2O).')
print(f"  N2O P_sat drop: {hist['p_feed_ox'][0]/1e6:.3f} -> {hist['p_feed_ox'][-1]/1e6:.3f} MPa  "
      f"({hist['n2o_temp'][0]-273.15:.1f} -> {hist['n2o_temp'][-1]-273.15:.1f} °C)")

# ══════════════════════════════════════════════════════════════════════════════
#  c* root-cause analysis
# ══════════════════════════════════════════════════════════════════════════════
from PistonLiquidRocket import combustion_props

A_t    = np.pi / 4 * (1.0 * IN) ** 2

# rocketCEA at the rounded O/F=2.00 that HalfCatSim's VLOOKUP uses (rounds to nearest 0.25)
of_hcs = 2.00   # HalfCatSim rounds O/F=2.10 → 2.00 for table lookup
Tc_cea,  gam_cea,  mw_cea,  cstar_cea  = combustion_props(E85_HCS, of_hcs, 250*6894.76, 4.0)
Tc_sim,  gam_sim,  mw_sim,  cstar_sim  = combustion_props(E85_HCS, 2.10,   REF['pc_mpa']*1e6, 4.0)

cstar_hcs_implied = REF['pc_mpa'] * 1e6 * A_t / REF['mdot_kgs']   # c*_real = Pc×At/mdot
cstar_hcs_table   = 1316.5   # m/s — directly read from HalfCatSim Fuel sheet, E85 c* at Pc=250 psi, O/F=2.00
cstar_ideal_hcs   = cstar_hcs_implied / 0.60
Tc_hcs_table      = 1734.0   # K — HalfCatSim Fuel sheet T0 at Pc=250 psi, O/F=2.00
ratio_cstar       = cstar_cea / cstar_hcs_table
ratio_Tc          = Tc_cea   / Tc_hcs_table
eta_needed        = 0.60 / ratio_cstar

W2 = 72
print()
print('=' * W2)
print('  c* Root-Cause: rocketCEA vs HalfCatSim CEA Table  (E85 / N2O)')
print('=' * W2)
print(f"  rocketCEA  (O/F=2.00, Pc=250 psi, full equilibrium):")
print(f"    Tc      = {Tc_cea:.0f} K       MW = {mw_cea:.2f} g/mol   γ = {gam_cea:.4f}")
print(f"    c*_ideal = {cstar_cea:.1f} m/s")
print()
print(f"  HalfCatSim Fuel sheet  (O/F=2.00, Pc=250 psi):")
print(f"    Tc      = {Tc_hcs_table:.0f} K  (51% lower than rocketCEA)")
print(f"    c*_ideal = {cstar_hcs_table:.1f} m/s  (24% lower → from table)")
print(f"    c*_real  = {cstar_hcs_table*0.60:.1f} m/s  (×η=0.60)")
print(f"    c*_real  = {cstar_hcs_implied:.1f} m/s  (implied: Pc×At/ṁ)")
print()
print(f"  CEA / HCS table ratio:  c* = {ratio_cstar:.3f}   Tc = {ratio_Tc:.3f}")
print()
print(f"  Variation with O/F (Tc_CEA / Tc_HCS at Pc=250 psi):")
from PistonLiquidRocket import combustion_props as _cp
import openpyxl as _xl, sys as _sys
try:
    _wb  = _xl.load_workbook(
        r'A:\Projects\Rockets\Liquid\Mojave Sphinx - Release 01\Simulation'
        r'\HalfCatSim_v1.3.8_MojaveSphinx - R01.xlsx', data_only=True)
    _ws  = _wb['Fuel']
    _hdr = [_ws.cell(row=4, column=c).value for c in range(478, 487)]
    _row = [_ws.cell(row=9, column=c).value for c in range(478, 487)]  # Pc=250 row
    print(f"  {'O/F':>5}  {'Tc_CEA':>8}  {'Tc_HCS':>8}  {'ratio':>6}  {'c*_CEA':>8}  {'c*_HCS':>8}")
    for of_v, Tc_h in zip(_hdr[1:], _row[1:]):
        if of_v is None or Tc_h is None: continue
        Tc_c, _, _, cs_c = _cp(E85_HCS, float(of_v), 250*6894.76, 4.0)
        print(f"  {of_v:>5.2f}  {Tc_c:>8.0f}  {Tc_h:>8.0f}  {Tc_c/Tc_h:>6.3f}", end='')
        # c* from table at this OF (col 479 onward at row 75 for Pc=250)
        # offset: OF 0.25..2.50 = cols 479..488, index = (of_v-0.25)/0.25
        idx   = round((float(of_v)-0.25)/0.25)
        cs_h  = _ws.cell(row=75, column=479+idx).value
        if cs_h: print(f"  {cs_c:>8.1f}  {cs_h:>8.1f}")
        else:    print()
except Exception as e:
    print(f"  (could not read spreadsheet: {e})")
print()
print(f"  KEY FINDINGS (from spreadsheet cell-level investigation):")
print(f"  1. HCS Fuel sheet confirmed: T0=1733.5 K, c*=1316.5 m/s, MW=18.75, γ=1.290")
print(f"     at O/F=2.00, Pc=250 psi for E85.  MW and γ match rocketCEA closely.")
print(f"     Despite near-identical gas composition, T0 is 51% lower → c* is 24% lower.")
print(f"     (c* ∝ √T0, so T0 ratio 1.51 → c* ratio 1.23 ≈ 1.24 measured)")
print()
print(f"  2. N2O enthalpy sensitivity (h°f = +82.05 kJ/mol → 0 kJ/mol 'cold N2O'):")
print(f"     Removing N2O decomp energy lowers T0 to 2052 K at O/F=2.00 — still")
print(f"     18% above HCS 1733 K.  Cold N2O matches HCS c* exactly at O/F=3.00,")
print(f"     diverges at fuel-rich conditions (O/F<3).  No single enthalpy offset")
print(f"     or inlet temperature explains the discrepancy across all O/F values.")
print()
print(f"  3. The HCS table is internally consistent: T0+MW+γ reproduce c* exactly.")
print(f"     η_c*=0.60 in HCS was calibrated so HCS actual c* = measured c* (790 m/s).")
print()
print(f"  4. rocketCEA ideal c* is 24% higher → applying same η=0.60 gives")
print(f"     c*_actual=982 m/s vs measured 790 m/s → overestimates Pc/Thrust/Isp by ~24%.")
print()
print(f"  CORRECTION: To match measured data when using rocketCEA:")
print(f"    cstar_eta = {eta_needed:.3f}  (= 0.60 × 1316.5/1635.5, O/F=2.00 reference)")
print(f"  Note: ratio varies with O/F (converges at O/F≈5).  Use η=0.483 as conservative")
print(f"  fuel-rich estimate; re-evaluate once hot-fire data is available.")
print('=' * W2)
print()
